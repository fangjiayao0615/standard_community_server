# -*- coding:utf-8 -*-
import ujson
from openpyxl import load_workbook
import requests

from cores.const import const_user
from cores.utils import logger


def batch_create_posts():
    """
    批量创建帖子
    :return: 
    """
    xls_file = "/Users/fangjiayao/Documents/community/data/all_data.xlsx"
    # sheet_name = "posts_funny"
    sheet_name = "posts_knowledge"
    wb = load_workbook(xls_file)
    ws = wb[sheet_name]

    # 上传图片
    upload_url = 'http://101.200.169.22/app/resource/upload_img'
    # 创建帖子
    create_post_url = 'http://101.200.169.22/audit/post/create'

    for i in range(1, ws.max_row):
        try:
            created = ws[i+1][0].value
            uid = ws[i+1][1].value.strip()
            tids = [tid.strip() for tid in ws[i+1][2].value.split(',')]
            ptype = int(ws[i+1][3].value)
            text = ws[i+1][4].value.strip()
            local_imgs = ws[i+1][5].value.split('\n') if ws[i+1][5].value else []
            # 已经创建用户
            if created:
                logger.info('post is already created uid=%s text=%s' % (uid, text))
                continue
            # 上传头像
            raw_imgs = []
            for local_img in local_imgs:
                local_img = local_img.strip()
                fl = open(local_img, 'rb')
                multiple_files = [
                    ('uploadedfile', (local_img.split("/")[-1], fl, 'image/jpeg'))
                ]
                upload_resp = requests.post(upload_url, files=multiple_files)
                upload_content = ujson.loads(upload_resp.content)
                print('upload image, resp %s' % upload_content)
                raw_imgs.append({
                    'url': upload_content['data']['url'],
                    'type': upload_content['data']['type'],
                    'w': upload_content['data']['w'],
                    'h': upload_content['data']['h']
                })
            # 创建用户
            data = {
                'session': const_user.SPIDER_ROB_ADMIN_SID,
                'uid': uid,
                'ptype': ptype,
                'text': text,
                'title': '',
                'raw_imgs': raw_imgs,
                'tids': tids,
                'raw_articles': [],
            }
            resp = requests.post(create_post_url, data=ujson.dumps(data))
            resp_content = ujson.loads(resp.content)
            print('create post, resp %s, text %s' % (resp_content, text))
        except:
            print('create post, failed, text %s' % text)
            continue


if __name__ == '__main__':
    batch_create_posts()
