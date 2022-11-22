# -*- coding:utf-8 -*-
import ujson
from openpyxl import load_workbook
import requests

from cores.const import const_user
from cores.utils import logger


def batch_create_users():
    """
    批量创建测试用户
    :return: 
    """
    xls_file = "/Users/fangjiayao/Documents/community/data/all_data.xlsx"
    sheet_name = "users_knowledge"
    wb = load_workbook(xls_file)
    ws = wb[sheet_name]

    # 上传头像
    upload_url = 'http://101.200.169.22/app/resource/upload_img'
    # 创建用户
    create_user_url = 'http://101.200.169.22/audit/user/create'

    for i in range(1, ws.max_row):
        uid = ws[i+1][0].value
        name = ws[i+1][1].value.strip()
        nick = ws[i+1][2].value.strip()
        sign = ws[i+1][3].value.strip()
        utypes = [int(utype.strip()) for utype in ws[i+1][4].value.split(',')]
        status = int(ws[i+1][5].value)
        local_avatar = ws[i+1][6].value.strip()
        # 已经创建用户
        if uid:
            logger.info('uid is already created %s %s %s' % (uid, name, nick))
            continue
        # 上传头像
        fl = open(local_avatar, 'rb')
        multiple_files = [
            ('uploadedfile', (local_avatar.split("/")[-1], fl, 'image/jpeg'))
        ]
        upload_resp = requests.post(upload_url, files=multiple_files)
        upload_content = ujson.loads(upload_resp.content)
        print('upload avatar, resp %s' % upload_content)
        # 创建用户
        data = {
            'session': const_user.SPIDER_ROB_ADMIN_SID,
            'name': name,
            'nick': nick,
            'status': status,
            'sign': sign,
            'raw_avatar': {
                'url': upload_content['data']['url'],
                'type': upload_content['data']['type'],
                'w': upload_content['data']['w'],
                'h': upload_content['data']['h']
            },
            'utypes': utypes,
        }
        resp = requests.post(create_user_url, data=ujson.dumps(data))
        resp_content = ujson.loads(resp.content)
        print('create user, resp %s' % resp_content)


if __name__ == '__main__':
    batch_create_users()
